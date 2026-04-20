"""
Admin router: platform-wide management, analytics, and exports.
"""
import csv
import io
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from geoalchemy2.shape import from_shape
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from shapely.geometry import Point
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session, aliased

from database import get_db
from models.bin import Bin, BinStatus
from models.collection import Collection
from models.report import BinReport, SHGReport
from models.route import Route, RouteStop
from models.user import User, UserRole
from models.zone import Zone
from schemas.admin import AdminSettings
from schemas.auth import UserRead
from schemas.bin import BinCreate, BinRead, BinUpdate
from schemas.user import UserCreate, UserUpdate
from services.auth_service import hash_password, require_role
from services.report_utils import normalize_bin_status
from services.route_optimizer import create_route_for_zone
from models.settings import SystemSettings
from models.recycler import Recycler, RecyclerBid, BidStatus
from services.notification_service import save_and_send_notification

router = APIRouter(prefix="/api/admin", tags=["Admin"])

HEADER_FILL = PatternFill(fill_type="solid", fgColor="2D6A4F")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _get_settings(db: Session) -> AdminSettings:
    """Read settings from DB, falling back to app config defaults."""
    from config import settings as app_settings

    rows = {row.key: row.value for row in db.query(SystemSettings).all()}

    def _val(db_key: str, default_attr: str) -> float:
        raw = rows.get(db_key)
        if raw is not None:
            return float(raw)
        return float(getattr(app_settings, default_attr))

    return AdminSettings(
        geofence_radius_meters=_val("geofence_radius_meters", "GEOFENCE_RADIUS_METERS"),
        ai_confidence_threshold=_val("ai_confidence_threshold", "AI_CONFIDENCE_THRESHOLD"),
        bin_collection_threshold_percent=_val("bin_collection_threshold_percent", "BIN_COLLECTION_THRESHOLD_PERCENT"),
        spam_window_minutes=_val("spam_window_minutes", "SPAM_WINDOW_MINUTES"),
        default_truck_capacity_kg=_val("default_truck_capacity_kg", "DEFAULT_TRUCK_CAPACITY_KG"),
    )


def _format_datetime(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M") if value else ""


def _format_date(value: date | None) -> str:
    return value.isoformat() if value else ""


def _collection_history_rows(db: Session) -> tuple[list[str], list[list[object]]]:
    rows = (
        db.query(Collection, User, Bin, Zone)
        .join(User, Collection.collector_id == User.id)
        .join(Bin, Collection.bin_id == Bin.id)
        .join(Zone, Bin.zone_id == Zone.id)
        .order_by(Collection.collected_at.desc(), Collection.id.desc())
        .all()
    )

    data = [
        [
            _format_date(collection.collected_at.date() if collection.collected_at else None),
            collector.full_name,
            bin_obj.label,
            zone.name,
            round(float(collection.kg_collected or 0), 2),
            collection.collected_at.strftime("%H:%M") if collection.collected_at else "",
            collection.notes or "",
        ]
        for collection, collector, bin_obj, zone in rows
    ]
    headers = ["Date", "Collector Name", "Bin Name", "Zone", "KG Collected", "Time", "Notes"]
    return headers, data


def _user_performance_rows(db: Session) -> tuple[list[str], list[list[object]]]:
    collectors = (
        db.query(User, Zone)
        .outerjoin(Zone, User.zone_id == Zone.id)
        .filter(User.role == UserRole.collector)
        .order_by(User.full_name.asc())
        .all()
    )

    rows: list[list[object]] = []
    for collector, zone in collectors:
        collections = (
            db.query(Collection)
            .filter(Collection.collector_id == collector.id)
            .order_by(Collection.collected_at.desc())
            .all()
        )
        total_collections = len(collections)
        total_kg = sum(float(collection.kg_collected or 0) for collection in collections)
        active_days = {
            collection.collected_at.date()
            for collection in collections
            if collection.collected_at
        }
        last_active = collections[0].collected_at if collections else None
        avg_kg_per_day = total_kg / len(active_days) if active_days else 0.0
        rows.append(
            [
                collector.full_name,
                collector.email,
                zone.name if zone else "",
                total_collections,
                round(total_kg, 2),
                round(avg_kg_per_day, 2),
                _format_datetime(last_active),
            ]
        )

    headers = [
        "Collector Name",
        "Email",
        "Zone",
        "Total Collections",
        "Total KG",
        "Avg KG/Day",
        "Last Active",
    ]
    return headers, rows


def _zone_summary_rows(db: Session) -> tuple[list[str], list[list[object]]]:
    month_start = date.today().replace(day=1)
    zones = db.query(Zone).order_by(Zone.name.asc()).all()
    rows: list[list[object]] = []

    for zone in zones:
        zone_bins = db.query(Bin).filter(Bin.zone_id == zone.id).all()
        total_bins = len(zone_bins)
        active_bins = sum(
            1
            for bin_obj in zone_bins
            if normalize_bin_status(bin_obj.status, bin_obj.fill_level) != "inactive"
        )
        collections_this_month = (
            db.query(func.count(Collection.id))
            .join(Bin, Collection.bin_id == Bin.id)
            .filter(Bin.zone_id == zone.id, func.date(Collection.collected_at) >= month_start)
            .scalar()
            or 0
        )
        total_kg_this_month = (
            db.query(func.coalesce(func.sum(Collection.kg_collected), 0.0))
            .join(Bin, Collection.bin_id == Bin.id)
            .filter(Bin.zone_id == zone.id, func.date(Collection.collected_at) >= month_start)
            .scalar()
            or 0.0
        )
        active_collectors = (
            db.query(func.count(User.id))
            .filter(
                User.zone_id == zone.id,
                User.role == UserRole.collector,
                User.is_active.is_(True),
            )
            .scalar()
            or 0
        )
        rows.append(
            [
                zone.name,
                total_bins,
                active_bins,
                int(collections_this_month),
                round(float(total_kg_this_month), 2),
                int(active_collectors),
            ]
        )

    headers = [
        "Zone Name",
        "Total Bins",
        "Active Bins",
        "Collections This Month",
        "Total KG This Month",
        "Active Collectors",
    ]
    return headers, rows


def _ai_analysis_rows(db: Session) -> tuple[list[str], list[list[object]]]:
    verifier_alias = aliased(User)
    rows = (
        db.query(BinReport, Bin, Zone, verifier_alias)
        .join(Bin, BinReport.bin_id == Bin.id)
        .join(Zone, Bin.zone_id == Zone.id)
        .outerjoin(verifier_alias, BinReport.verified_by == verifier_alias.id)
        .order_by(BinReport.created_at.desc(), BinReport.id.desc())
        .all()
    )

    data = [
        [
            _format_date(report.created_at.date() if report.created_at else None),
            bin_obj.label,
            zone.name,
            report.fill_level or 0,
            report.waste_type or "",
            report.urgency or "",
            round(float(report.ai_confidence or 0), 2),
            report.ai_observations or "",
            verifier.full_name if verifier else "",
            report.status,
        ]
        for report, bin_obj, zone, verifier in rows
    ]
    headers = [
        "Date",
        "Bin Name",
        "Zone",
        "Reported Fill Level",
        "Waste Type",
        "Urgency",
        "AI Confidence",
        "AI Observations",
        "Verified By",
        "Status",
    ]
    return headers, data


def _bin_status_rows(db: Session) -> tuple[list[str], list[list[object]]]:
    collection_stats = {
        row.bin_id: row
        for row in (
            db.query(
                Collection.bin_id.label("bin_id"),
                func.max(Collection.collected_at).label("last_collected"),
                func.count(Collection.id).label("total_collections"),
            )
            .group_by(Collection.bin_id)
            .all()
        )
    }
    rows = (
        db.query(Bin, Zone)
        .join(Zone, Bin.zone_id == Zone.id)
        .order_by(Zone.name.asc(), Bin.label.asc())
        .all()
    )

    data = []
    for bin_obj, zone in rows:
        stats = collection_stats.get(bin_obj.id)
        last_collected = stats.last_collected if stats and stats.last_collected else bin_obj.last_collected
        total_collections = int(stats.total_collections) if stats else 0
        data.append(
            [
                bin_obj.label,
                zone.name,
                bin_obj.address or "",
                int(bin_obj.fill_level or 0),
                normalize_bin_status(bin_obj.status, bin_obj.fill_level),
                _format_datetime(last_collected),
                total_collections,
            ]
        )

    headers = [
        "Bin Name",
        "Zone",
        "Address",
        "Current Fill %",
        "Status",
        "Last Collected",
        "Total Collections",
    ]
    return headers, data


def _build_dashboard_data(db: Session) -> dict:
    total_plastic_kg = float(
        db.query(func.coalesce(func.sum(Collection.kg_collected), 0.0)).scalar() or 0.0
    )
    fuel_saved_liters = round(total_plastic_kg * 0.3, 2)
    co2_reduced_kg = round(total_plastic_kg * 2.5, 2)

    bins = db.query(Bin).all()
    active_bins = 0
    distribution = {"empty": 0, "high": 0, "full": 0, "inactive": 0}
    for bin_obj in bins:
        normalized = normalize_bin_status(bin_obj.status, bin_obj.fill_level)
        if normalized != "inactive":
            active_bins += 1
        if normalized in distribution:
            distribution[normalized] += 1

    total_users = db.query(func.count(User.id)).scalar() or 0
    pending_reports = (
        db.query(func.count(BinReport.id))
        .filter(BinReport.status == "pending")
        .scalar()
        or 0
    )

    daily_collections = []
    for day_offset in range(6, -1, -1):
        current_day = date.today() - timedelta(days=day_offset)
        kg = (
            db.query(func.coalesce(func.sum(Collection.kg_collected), 0.0))
            .filter(func.date(Collection.collected_at) == current_day)
            .scalar()
            or 0.0
        )
        count = (
            db.query(func.count(Collection.id))
            .filter(func.date(Collection.collected_at) == current_day)
            .scalar()
            or 0
        )
        daily_collections.append(
            {
                "date": current_day.strftime("%a"),
                "kg": round(float(kg), 2),
                "count": int(count),
            }
        )

    month_start = date.today().replace(day=1)
    zone_performance = []
    zones = db.query(Zone).order_by(Zone.name.asc()).all()
    for zone in zones:
        total_bins = db.query(func.count(Bin.id)).filter(Bin.zone_id == zone.id).scalar() or 0
        collections_this_month = (
            db.query(func.count(Collection.id))
            .join(Bin, Collection.bin_id == Bin.id)
            .filter(Bin.zone_id == zone.id, func.date(Collection.collected_at) >= month_start)
            .scalar()
            or 0
        )
        kg_this_month = (
            db.query(func.coalesce(func.sum(Collection.kg_collected), 0.0))
            .join(Bin, Collection.bin_id == Bin.id)
            .filter(Bin.zone_id == zone.id, func.date(Collection.collected_at) >= month_start)
            .scalar()
            or 0.0
        )
        zone_performance.append(
            {
                "zone_name": zone.name,
                "total_bins": int(total_bins),
                "collections_this_month": int(collections_this_month),
                "kg_this_month": round(float(kg_this_month), 2),
            }
        )

    # Waste type distribution from verified/AI-analysed bin reports
    waste_type_rows = (
        db.query(
            func.coalesce(func.nullif(func.trim(BinReport.waste_type), ""), "unknown").label("wtype"),
            func.count(BinReport.id).label("cnt"),
        )
        .group_by(
            func.coalesce(func.nullif(func.trim(BinReport.waste_type), ""), "unknown")
        )
        .order_by(func.count(BinReport.id).desc())
        .all()
    )
    waste_type_distribution = [
        {"name": row.wtype.capitalize(), "value": int(row.cnt)}
        for row in waste_type_rows
    ]

    return {
        "total_plastic_kg": round(total_plastic_kg, 2),
        "fuel_saved_liters": fuel_saved_liters,
        "co2_reduced_kg": co2_reduced_kg,
        "active_bins": int(active_bins),
        "total_users": int(total_users),
        "pending_reports": int(pending_reports),
        "daily_collections": daily_collections,
        "zone_performance": zone_performance,
        "bin_status_distribution": distribution,
        "waste_type_distribution": waste_type_distribution,
    }


def _write_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in rows:
        worksheet.append(row)

    for column in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column]
        max_length = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = max_length + 2


# Dashboard
@router.get("/dashboard")
def get_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    return _build_dashboard_data(db)


# Zones
@router.get("/zones")
def list_zones(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    zones = db.query(Zone).order_by(Zone.name.asc()).all()
    return [
        {
            "id": z.id,
            "name": z.name,
            "description": z.description,
            "center_lat": z.center_lat,
            "center_lng": z.center_lng,
            "radius_km": z.radius_km,
            "depot_lat": z.depot_lat,
            "depot_lng": z.depot_lng,
            "depot_address": z.depot_address,
        }
        for z in zones
    ]

@router.post("/zones", status_code=201)
def create_zone(
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    try:
        zone = Zone(
            name=data.get('name', 'New Zone'),
            description=data.get('description', ''),
        )
        db.add(zone)
        db.commit()
        db.refresh(zone)
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Zone created: {zone.id} - {zone.name}")
        return {
            "id": zone.id,
            "name": zone.name,
            "description": zone.description,
            "center_lat": zone.center_lat,
            "center_lng": zone.center_lng,
            "radius_km": zone.radius_km,
            "depot_lat": zone.depot_lat,
            "depot_lng": zone.depot_lng,
            "depot_address": zone.depot_address,
        }
    except Exception as e:
        db.rollback()
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Zone create failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create zone: {str(e)}"
        )


@router.put("/zones/{zone_id}")
def update_zone(
    zone_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    zone = db.query(Zone).filter(Zone.id == zone_id).first()
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    
    try:
        # Only update fields that exist on the model
        if 'name' in data and data['name']:
            zone.name = data['name']
        if hasattr(zone, 'description') and 'description' in data:
            zone.description = data['description']
        if hasattr(zone, 'center_lat') and 'center_lat' in data:
            zone.center_lat = data['center_lat']
        if hasattr(zone, 'center_lng') and 'center_lng' in data:
            zone.center_lng = data['center_lng']
        if hasattr(zone, 'radius_km') and 'radius_km' in data:
            zone.radius_km = data['radius_km']
        if hasattr(zone, 'depot_lat') and 'depot_lat' in data:
            zone.depot_lat = data['depot_lat']
        if hasattr(zone, 'depot_lng') and 'depot_lng' in data:
            zone.depot_lng = data['depot_lng']
        if hasattr(zone, 'depot_address') and 'depot_address' in data:
            zone.depot_address = data['depot_address']
        
        db.commit()
        db.refresh(zone)
        # Safely return attributes
        return {
            "id": zone.id, "name": zone.name, "description": getattr(zone, 'description', None),
            "center_lat": getattr(zone, 'center_lat', None), "center_lng": getattr(zone, 'center_lng', None),
            "radius_km": getattr(zone, 'radius_km', None), "depot_lat": getattr(zone, 'depot_lat', None),
            "depot_lng": getattr(zone, 'depot_lng', None), "depot_address": getattr(zone, 'depot_address', None),
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update zone: {str(e)}")


@router.delete("/zones/{zone_id}")
def delete_zone(
    zone_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    zone = db.query(Zone).filter(Zone.id == zone_id).first()
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    
    try:
        # Find another zone for reassignment
        other_zone = db.query(Zone).filter(Zone.id != zone_id).first()
        
        # 1. Handle Bins (NOT NULL)
        bins_count = db.query(Bin).filter(Bin.zone_id == zone_id).count()
        if bins_count > 0:
            if not other_zone:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot delete last zone: {bins_count} bins assigned."
                )
            db.query(Bin).filter(Bin.zone_id == zone_id).update(
                {"zone_id": other_zone.id}, synchronize_session=False
            )
            
        # 2. Handle Recyclers (NOT NULL)
        from models.recycler import Recycler
        recyclers_count = db.query(Recycler).filter(Recycler.zone_id == zone_id).count()
        if recyclers_count > 0:
            if not other_zone:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot delete last zone: {recyclers_count} recyclers assigned."
                )
            db.query(Recycler).filter(Recycler.zone_id == zone_id).update(
                {"zone_id": other_zone.id}, synchronize_session=False
            )

        # 3. Handle SHGReports (NOT NULL)
        from models.report import SHGReport
        shg_reports_count = db.query(SHGReport).filter(SHGReport.zone_id == zone_id).count()
        if shg_reports_count > 0:
            if not other_zone:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot delete last zone: {shg_reports_count} SHG reports mapped here."
                )
            db.query(SHGReport).filter(SHGReport.zone_id == zone_id).update(
                {"zone_id": other_zone.id}, synchronize_session=False
            )
        
        # 4. Handle Users (Nullable - set to None)
        db.query(User).filter(User.zone_id == zone_id).update(
            {"zone_id": None}, synchronize_session=False
        )
        
        # 5. Handle Routes (Delete)
        from models.route import Route
        db.query(Route).filter(Route.zone_id == zone_id).delete(synchronize_session=False)
        
        # 6. Delete zone itself
        db.flush()
        db.delete(zone)
        db.commit()
        
        msg = "Zone deleted successfully"
        reassignments = []
        if bins_count > 0 and other_zone: reassignments.append(f"{bins_count} bins")
        if recyclers_count > 0 and other_zone: reassignments.append(f"{recyclers_count} recyclers")
        if shg_reports_count > 0 and other_zone: reassignments.append(f"{shg_reports_count} SHG reports")
        
        if reassignments and other_zone:
            msg += f" ({', '.join(reassignments)} moved to {other_zone.name})"
            
        return {"message": msg}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to delete zone: {str(e)}"
        )


# User Management
@router.get("/users", response_model=list[UserRead])
def list_users(
    role: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    return query.all()


@router.post("/users", response_model=UserRead)
def create_user(
    data: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    existing = db.query(User).filter(User.email == data.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Validate phone_number uniqueness
    if data.phone_number:
        phone_exists = db.query(User).filter(User.phone_number == data.phone_number).first()
        if phone_exists:
            raise HTTPException(status_code=400, detail="Phone number already registered")

    # Validate zone exists
    if data.zone_id:
        zone = db.query(Zone).filter(Zone.id == data.zone_id).first()
        if not zone:
            raise HTTPException(
                status_code=400,
                detail=f"Zone with id {data.zone_id} not found. Please select a valid zone."
            )

    user = User(
        full_name=data.name,
        email=data.email,
        hashed_password=hash_password(data.password),
        role=data.role,
        zone_id=data.zone_id,
        phone_number=data.phone_number,
    )
    try:
        db.add(user)
        db.commit()
        db.refresh(user)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="A user with this email or phone number already exists")
    return user


@router.put("/users/{user_id}", response_model=UserRead)
def update_user(
    user_id: int,
    data: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if data.full_name is not None:
        user.full_name = data.full_name
    if data.phone is not None:
        user.phone = data.phone
    if data.phone_number is not None:
        # Validate phone_number uniqueness (exclude current user)
        if data.phone_number:
            phone_exists = db.query(User).filter(
                User.phone_number == data.phone_number,
                User.id != user_id
            ).first()
            if phone_exists:
                raise HTTPException(status_code=400, detail="Phone number already registered to another user")
        user.phone_number = data.phone_number
    if data.is_active is not None:
        user.is_active = data.is_active
    if data.zone_id is not None:
        user.zone_id = data.zone_id

    try:
        db.commit()
        db.refresh(user)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="A user with this phone number already exists")
    return user


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    db.query(BinReport).filter(BinReport.verified_by == user_id).update({"verified_by": None}, synchronize_session=False)
    db.query(SHGReport).filter(SHGReport.shg_user_id == user_id).delete(synchronize_session=False)
    db.query(Collection).filter(Collection.collector_id == user_id).delete(synchronize_session=False)
    db.query(Route).filter(Route.collector_id == user_id).delete(synchronize_session=False)
    
    db.delete(user)
    db.commit()
    return {"message": "User deleted successfully"}


# Bin Management
@router.get("/bins", response_model=list[BinRead])
def list_bins(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    return db.query(Bin).all()


@router.post("/bins", response_model=BinRead)
def create_bin(
    data: BinCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    bin_obj = Bin(
        label=data.label,
        location=from_shape(Point(data.longitude, data.latitude), srid=4326),
        latitude=data.latitude,
        longitude=data.longitude,
        address=data.address,
        zone_id=data.zone_id,
        capacity_kg=data.capacity_kg,
        status=BinStatus.empty,
        fill_level=0,
    )
    db.add(bin_obj)
    db.commit()
    db.refresh(bin_obj)
    return bin_obj


@router.put("/bins/{bin_id}", response_model=BinRead)
def update_bin(
    bin_id: int,
    data: BinUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    bin_obj = db.query(Bin).filter(Bin.id == bin_id).first()
    if not bin_obj:
        raise HTTPException(status_code=404, detail="Bin not found")

    if data.label is not None:
        bin_obj.label = data.label
    if data.address is not None:
        bin_obj.address = data.address
    if data.status is not None:
        bin_obj.status = data.status
    if data.fill_level is not None:
        bin_obj.fill_level = data.fill_level
    if data.zone_id is not None:
        bin_obj.zone_id = data.zone_id
    if data.capacity_kg is not None:
        bin_obj.capacity_kg = data.capacity_kg

    db.commit()
    db.refresh(bin_obj)
    return bin_obj


@router.delete("/bins/{bin_id}")
def delete_bin(
    bin_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    bin_obj = db.query(Bin).filter(Bin.id == bin_id).first()
    if not bin_obj:
        raise HTTPException(status_code=404, detail="Bin not found")

    # Delete related records first to avoid FK constraint errors
    db.query(BinReport).filter(BinReport.bin_id == bin_id).delete(synchronize_session=False)
    db.query(Collection).filter(Collection.bin_id == bin_id).delete(synchronize_session=False)
    db.query(RouteStop).filter(RouteStop.bin_id == bin_id).delete(synchronize_session=False)
    db.query(SHGReport).filter(SHGReport.bin_id == bin_id).delete(synchronize_session=False)

    db.delete(bin_obj)
    db.commit()
    return {"message": "Bin deleted successfully"}


# Reports
@router.get("/reports")
def list_reports(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    headers, rows = _ai_analysis_rows(db)
    return [dict(zip(headers, row)) for row in rows]


# Analytics
@router.get("/analytics")
def get_analytics(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    dashboard = _build_dashboard_data(db)
    summary = {
        key: value
        for key, value in dashboard.items()
        if key not in {"daily_collections", "zone_performance", "bin_status_distribution"}
    }
    return {
        "summary": summary,
        "daily_collections": dashboard["daily_collections"],
        "zone_performance": dashboard["zone_performance"],
        "bin_status_distribution": [
            {"status": status, "count": count}
            for status, count in dashboard["bin_status_distribution"].items()
        ],
    }


# Route Generation
@router.post("/routes/generate")
def generate_routes(
    zone_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    zone_ids = [zone_id] if zone_id is not None else [zone.id for zone in db.query(Zone).all()]
    if zone_id is not None and not zone_ids:
        raise HTTPException(status_code=404, detail="Zone not found")

    generated_routes = []
    for target_zone_id in zone_ids:
        try:
            result = create_route_for_zone(
                db,
                zone_id=target_zone_id,
                collection_threshold_percent=60,
                route_name_prefix="AI Route",
            )
            if result["route_optimized"]:
                db.commit()
                generated_routes.append(result)

                # Notify the assigned collector about new route
                try:
                    route_obj = db.query(Route).filter(Route.id == result["route_id"]).first()
                    if route_obj and route_obj.collector_id:
                        save_and_send_notification(
                            db=db,
                            user_id=route_obj.collector_id,
                            title="New Route Assigned 🚛",
                            body=f"You have a new route with {result['stops_count']} bins in {result['zone_name']} ({result['total_distance_km']} km).",
                            data={"type": "route_assigned", "route_id": result["route_id"]},
                        )
                        db.commit()
                except Exception:
                    pass  # Never block route generation
            else:
                db.rollback()
                if zone_id is not None:
                    return {"message": result["message"], "route": None}
        except ValueError as exc:
            db.rollback()
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except SQLAlchemyError as exc:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to save route: {exc}") from exc

    if not generated_routes:
        return {"message": "No bins require collection at this time", "route": None}

    if zone_id is not None:
        route_data = generated_routes[0]
        route_id = route_data["route_id"]

        # Fetch the saved route stops with bin coordinates
        from models.route import RouteStop
        from models.bin import Bin
        stops = (
            db.query(RouteStop, Bin)
            .join(Bin, RouteStop.bin_id == Bin.id)
            .filter(RouteStop.route_id == route_id)
            .order_by(RouteStop.sequence.asc())
            .all()
        )
        stops_data = [
            {
                "sequence": stop.sequence,
                "bin_id": bin_obj.id,
                "label": bin_obj.label,
                "address": bin_obj.address or "",
                "lat": bin_obj.latitude,
                "lng": bin_obj.longitude,
                "fill_level": bin_obj.fill_level,
                "status": getattr(bin_obj.status, "value", str(bin_obj.status)),
            }
            for stop, bin_obj in stops
        ]

        # Fetch collector name
        from models.route import Route
        route_obj = db.query(Route).filter(Route.id == route_id).first()
        collector_name = None
        if route_obj and route_obj.collector:
            collector_name = route_obj.collector.full_name

        return {
            "route_id": route_id,
            "total_distance_km": route_data["total_distance_km"],
            "estimated_duration_min": route_data["estimated_minutes"],
            "bins_count": route_data["stops_count"],
            "zone_name": route_data["zone_name"],
            "collector": collector_name,
            "stops": stops_data,
        }

    all_stops_data = []
    for route_data in generated_routes:
        route_id = route_data["route_id"]
        from models.route import RouteStop
        from models.bin import Bin
        stops = (
            db.query(RouteStop, Bin)
            .join(Bin, RouteStop.bin_id == Bin.id)
            .filter(RouteStop.route_id == route_id)
            .order_by(RouteStop.sequence.asc())
            .all()
        )
        for stop, bin_obj in stops:
            all_stops_data.append({
                "sequence": stop.sequence,
                "bin_id": bin_obj.id,
                "label": bin_obj.label,
                "address": bin_obj.address or "",
                "lat": bin_obj.latitude,
                "lng": bin_obj.longitude,
                "fill_level": bin_obj.fill_level,
                "status": getattr(bin_obj.status, "value", str(bin_obj.status)),
                "zone_name": route_data["zone_name"],
            })

    return {
        "message": "Routes generated successfully",
        "routes_created": generated_routes,
        "count": len(generated_routes),
        "stops": all_stops_data,
    }


# Exports
@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheet_builders = [
        ("Collection History", _collection_history_rows),
        ("User Performance", _user_performance_rows),
        ("Zone Summary", _zone_summary_rows),
        ("AI Analysis", _ai_analysis_rows),
        ("Bin Status", _bin_status_rows),
    ]
    for title, builder in sheet_builders:
        headers, rows = builder(db)
        _write_sheet(workbook, title, headers, rows)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"smartwaste_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/csv")
def export_csv(
    type: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    builders = {
        "collections": _collection_history_rows,
        "users": _user_performance_rows,
        "zones": _zone_summary_rows,
        "ai_analysis": _ai_analysis_rows,
    }
    if type not in builders:
        raise HTTPException(status_code=400, detail="Unsupported CSV export type")

    headers, rows = builders[type](db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)

    filename = f"smartwaste_{type}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Settings
@router.get("/settings", response_model=AdminSettings)
def get_settings(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    return _get_settings(db)


@router.put("/settings", response_model=AdminSettings)
def update_settings(
    data: AdminSettings,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    updates = {
        "geofence_radius_meters": data.geofence_radius_meters,
        "ai_confidence_threshold": data.ai_confidence_threshold,
        "bin_collection_threshold_percent": data.bin_collection_threshold_percent,
        "spam_window_minutes": data.spam_window_minutes,
        "default_truck_capacity_kg": data.default_truck_capacity_kg,
    }
    for key, value in updates.items():
        row = db.query(SystemSettings).filter(SystemSettings.key == key).first()
        if row:
            row.value = str(value)
        else:
            db.add(SystemSettings(key=key, value=str(value)))
    try:
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to save settings: {exc}") from exc
    return _get_settings(db)


# Recycler Management
@router.get("/recyclers")
def list_recyclers(
    zone_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    """List all recyclers, optionally filtered by zone."""
    query = db.query(Recycler)
    if zone_id is not None:
        query = query.filter(Recycler.zone_id == zone_id)
    recyclers = query.order_by(Recycler.name.asc()).all()
    return [
        {
            "id": r.id,
            "name": r.name,
            "contact_person": r.contact_person,
            "phone": r.phone,
            "email": r.email,
            "address": r.address,
            "latitude": r.latitude,
            "longitude": r.longitude,
            "accepted_types": r.accepted_types,
            "price_per_kg": r.price_per_kg,
            "min_quantity_kg": r.min_quantity_kg,
            "zone_id": r.zone_id,
            "is_active": r.is_active,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in recyclers
    ]


@router.get("/recyclers/stats")
def get_recycler_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    """Return platform-wide recycler statistics."""
    total_recyclers = db.query(func.count(Recycler.id)).filter(Recycler.is_active.is_(True)).scalar() or 0
    total_bids = db.query(func.count(RecyclerBid.id)).scalar() or 0
    pending_bids = db.query(func.count(RecyclerBid.id)).filter(RecyclerBid.status == BidStatus.pending).scalar() or 0
    completed_bids = db.query(func.count(RecyclerBid.id)).filter(RecyclerBid.status == BidStatus.completed).scalar() or 0
    completed_rows = db.query(RecyclerBid).filter(RecyclerBid.status == BidStatus.completed).all()
    total_kg = round(sum(b.quantity_kg for b in completed_rows), 2)
    total_value = round(sum(b.quantity_kg * b.offered_price_per_kg for b in completed_rows), 2)
    return {
        "total_recyclers": int(total_recyclers),
        "total_bids": int(total_bids),
        "pending_bids": int(pending_bids),
        "completed_bids": int(completed_bids),
        "total_kg_processed": total_kg,
        "total_value": total_value,
    }


@router.post("/recyclers")
def create_recycler(
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    """Create a new recycler."""
    recycler = Recycler(
        name=data["name"],
        contact_person=data["contact_person"],
        phone=data["phone"],
        email=data.get("email"),
        address=data["address"],
        latitude=float(data["latitude"]),
        longitude=float(data["longitude"]),
        accepted_types=data.get("accepted_types", []),
        price_per_kg=float(data.get("price_per_kg", 0.0)),
        min_quantity_kg=float(data.get("min_quantity_kg", 0.0)),
        zone_id=int(data["zone_id"]),
        is_active=data.get("is_active", True),
    )
    db.add(recycler)
    db.commit()
    db.refresh(recycler)
    return {"id": recycler.id, "name": recycler.name, "message": "Recycler created successfully"}


@router.put("/recyclers/{recycler_id}")
def update_recycler(
    recycler_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    """Update an existing recycler."""
    recycler = db.query(Recycler).filter(Recycler.id == recycler_id).first()
    if not recycler:
        raise HTTPException(status_code=404, detail="Recycler not found")
    updatable = ["name", "contact_person", "phone", "email", "address",
                 "accepted_types", "price_per_kg", "min_quantity_kg", "zone_id", "is_active"]
    for field in updatable:
        if field in data:
            setattr(recycler, field, data[field])
    if "latitude" in data:
        recycler.latitude = float(data["latitude"])
    if "longitude" in data:
        recycler.longitude = float(data["longitude"])
    db.commit()
    db.refresh(recycler)
    return {"id": recycler.id, "name": recycler.name, "message": "Recycler updated successfully"}


@router.delete("/recyclers/{recycler_id}")
def delete_recycler(
    recycler_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    """Delete a recycler and all associated bids."""
    recycler = db.query(Recycler).filter(Recycler.id == recycler_id).first()
    if not recycler:
        raise HTTPException(status_code=404, detail="Recycler not found")
    db.query(RecyclerBid).filter(RecyclerBid.recycler_id == recycler_id).delete(synchronize_session=False)
    db.delete(recycler)
    db.commit()
    return {"message": "Recycler deleted successfully"}

@router.get('/routes')
def get_admin_routes(db: Session = Depends(get_db), current_user: User = Depends(require_role('admin'))):
    "Get all routes for the admin dashboard"
    routes = db.query(Route).order_by(Route.created_at.desc()).all()
    return [
        {
            'id': r.id,
            'name': r.name,
            'date': _format_date(r.date),
            'status': r.status.value,
            'total_distance_km': float(r.total_distance_km) if r.total_distance_km else 0.0,
            'total_estimated_time_mins': r.estimated_duration_min,
            'collector_id': r.collector_id,
            'zone_id': r.zone_id,
            'created_at': r.created_at.isoformat() if r.created_at else None,
            'updated_at': r.updated_at.isoformat() if r.updated_at else None
        }
        for r in routes
    ]

@router.post("/reseed-zones")
async def reseed_zones(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    existing = db.query(Zone).count()
    if existing > 0:
        return {"message": f"zones already exist"}
    
    zones = [
        Zone(name="Zone 1 (North)", description="Northern collection area"),
        Zone(name="Zone 2 (South)", description="Southern collection area"),
        Zone(name="Zone 3 (East)", description="Eastern collection area"),
        Zone(name="Zone 4 (West)", description="Western collection area"),
    ]
    db.add_all(zones)
    db.commit()
    return {"message": "4 zones created successfully"}
